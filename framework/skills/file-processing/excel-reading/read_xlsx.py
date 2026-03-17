"""
xlsx 文件阅读理解工具

使用 openpyxl 库读取 Excel 文件，提取工作表内容并返回结构化数据。
"""

from typing import Optional, List, Dict, Any, Union
from pathlib import Path
import json

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("请安装 openpyxl: pip install openpyxl")


def read_xlsx(
    file_path: Union[str, Path],
    sheet_names: Optional[List[str]] = None,
    max_rows: Optional[int] = None
) -> Dict[str, Any]:
    """
    读取 xlsx 文件并返回结构化数据。
    
    Args:
        file_path: Excel 文件路径
        sheet_names: 指定要读取的工作表名称列表，默认读取所有工作表
        max_rows: 每个工作表最大读取行数，用于限制大文件读取，默认无限制
    
    Returns:
        包含文件信息的字典，结构如下：
        {
            "file_info": {
                "file_path": "文件路径",
                "total_sheets": 工作表总数,
                "read_sheets": 实际读取的工作表数量
            },
            "sheets": [
                {
                    "name": "工作表名称",
                    "rows": 总行数,
                    "columns": 总列数,
                    "data": [[单元格数据...], ...]
                },
                ...
            ]
        }
    
    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 文件格式错误或工作表不存在
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
        raise ValueError(f"不支持的文件格式: {file_path.suffix}，仅支持 .xlsx、.xlsm 和 .xls 文件")
    
    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"无法打开文件: {e}")
    
    result: Dict[str, Any] = {
        "file_info": {
            "file_path": str(file_path),
            "total_sheets": len(workbook.sheetnames),
            "read_sheets": 0
        },
        "sheets": []
    }
    
    target_sheets = sheet_names if sheet_names else workbook.sheetnames
    
    for sheet_name in target_sheets:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"工作表不存在: {sheet_name}")
        
        worksheet = workbook[sheet_name]
        
        sheet_data: Dict[str, Any] = {
            "name": sheet_name,
            "rows": 0,
            "columns": 0,
            "data": []
        }
        
        rows_data = []
        row_count = 0
        max_column = 0
        
        for row in worksheet.iter_rows(values_only=True):
            if max_rows is not None and row_count >= max_rows:
                break
            
            row_list = list(row) if row else []
            
            row_cleaned = []
            for cell in row_list:
                if cell is None:
                    row_cleaned.append("")
                elif isinstance(cell, (int, float, bool)):
                    row_cleaned.append(cell)
                else:
                    row_cleaned.append(str(cell))
            
            rows_data.append(row_cleaned)
            row_count += 1
            
            if row:
                current_max = len(row)
                if current_max > max_column:
                    max_column = current_max
        
        sheet_data["rows"] = row_count
        sheet_data["columns"] = max_column
        sheet_data["data"] = rows_data
        
        result["sheets"].append(sheet_data)
    
    result["file_info"]["read_sheets"] = len(result["sheets"])
    
    workbook.close()
    
    return result


def read_xlsx_as_markdown(
    file_path: Union[str, Path],
    sheet_names: Optional[List[str]] = None,
    max_rows: Optional[int] = None
) -> str:
    """
    读取 xlsx 文件并返回 Markdown 格式的表格。
    
    Args:
        file_path: Excel 文件路径
        sheet_names: 指定要读取的工作表名称列表
        max_rows: 每个工作表最大读取行数
    
    Returns:
        Markdown 格式的字符串
    """
    data = read_xlsx(file_path, sheet_names, max_rows)
    
    markdown_lines = []
    markdown_lines.append(f"# Excel 文件: {data['file_info']['file_path']}\n")
    markdown_lines.append(f"- 总工作表数: {data['file_info']['total_sheets']}")
    markdown_lines.append(f"- 读取工作表数: {data['file_info']['read_sheets']}\n")
    
    for sheet in data["sheets"]:
        markdown_lines.append(f"## 工作表: {sheet['name']}\n")
        markdown_lines.append(f"- 行数: {sheet['rows']}")
        markdown_lines.append(f"- 列数: {sheet['columns']}\n")
        
        if sheet["data"]:
            header = sheet["data"][0] if sheet["data"] else []
            markdown_lines.append("| " + " | ".join(str(cell) for cell in header) + " |")
            markdown_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            
            for row in sheet["data"][1:]:
                while len(row) < len(header):
                    row.append("")
                markdown_lines.append("| " + " | ".join(str(cell) for cell in row[:len(header)]) + " |")
        
        markdown_lines.append("")
    
    return "\n".join(markdown_lines)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python read_xlsx.py <文件路径> [工作表名称...] [--max-rows N]")
        print("示例: python read_xlsx.py data.xlsx Sheet1 Sheet2 --max-rows 100")
        sys.exit(1)
    
    file_path_arg = sys.argv[1]
    sheet_names_arg: Optional[List[str]] = None
    max_rows_arg: Optional[int] = None
    
    args = sys.argv[2:]
    if "--max-rows" in args:
        idx = args.index("--max-rows")
        if idx + 1 < len(args):
            max_rows_arg = int(args[idx + 1])
            args = args[:idx] + args[idx + 2:]
    
    if args:
        sheet_names_arg = args
    
    try:
        result = read_xlsx(file_path_arg, sheet_names_arg, max_rows_arg)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
